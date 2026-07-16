"""Behavioural tests for :mod:`bi_reports`.

The suite must pass with **stdlib only**: the CSV path is exercised fully, while
the Excel and PDF paths adapt to their optional engines — when the engine is
absent we assert the clean :class:`MissingDependencyError`; when it *is* present
(a fuller dev environment) we assert a real file is written. Detection uses a
plain ``importlib`` probe so the same test file behaves correctly either way.
"""

from __future__ import annotations

import csv
import importlib.util

import pytest

from bi_reports import (
    Dataset,
    ExportFormat,
    ExportResult,
    MissingDependencyError,
    ReportError,
    UnsupportedFormatError,
    export,
    export_csv,
    export_excel,
    export_pdf,
)

SAMPLE = Dataset(
    columns=("BRANCH", "TOTAL_SALES"),
    rows=(
        ("North", 120),
        ("South", 90),
        ("East", None),
    ),
)


def _installed(module: str) -> bool:
    """Return ``True`` if ``module`` can be imported without importing it."""
    return importlib.util.find_spec(module) is not None


# --------------------------------------------------------------------------- #
# CSV — always available
# --------------------------------------------------------------------------- #


def test_csv_writes_header_and_all_rows(tmp_path):
    path = tmp_path / "report.csv"
    export_csv(SAMPLE, str(path))

    with open(path, encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))

    assert rows[0] == ["BRANCH", "TOTAL_SALES"]
    assert rows[1] == ["North", "120"]
    assert len(rows) == 1 + SAMPLE.row_count  # header + data rows


def test_csv_null_becomes_empty_string(tmp_path):
    path = tmp_path / "report.csv"
    export_csv(SAMPLE, str(path))
    with open(path, encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    # The East row has a NULL sales value -> empty cell, not the text "None".
    assert rows[3] == ["East", ""]


def test_csv_result_metadata(tmp_path):
    path = tmp_path / "report.csv"
    result = export_csv(SAMPLE, str(path))
    assert isinstance(result, ExportResult)
    assert result.format == "csv"
    assert result.engine == "csv/stdlib"
    assert result.row_count == 3
    assert result.byte_size > 0
    assert result.byte_size == path.stat().st_size


def test_export_dispatches_to_csv(tmp_path):
    path = tmp_path / "via_dispatch.csv"
    result = export(SAMPLE, "csv", str(path))
    assert result.engine == "csv/stdlib"
    assert path.exists()


def test_export_accepts_enum_and_is_case_insensitive(tmp_path):
    p1 = tmp_path / "a.csv"
    p2 = tmp_path / "b.csv"
    assert export(SAMPLE, ExportFormat.CSV, str(p1)).format == "csv"
    assert export(SAMPLE, "CSV", str(p2)).format == "csv"


def test_empty_dataset_writes_header_only(tmp_path):
    empty = Dataset(columns=("A", "B"), rows=())
    path = tmp_path / "empty.csv"
    result = export_csv(empty, str(path))
    with open(path, encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows == [["A", "B"]]
    assert result.row_count == 0


# --------------------------------------------------------------------------- #
# Unsupported format
# --------------------------------------------------------------------------- #


def test_unsupported_format_raises(tmp_path):
    with pytest.raises(UnsupportedFormatError):
        export(SAMPLE, "html", str(tmp_path / "x.html"))


def test_unsupported_format_error_is_report_error():
    assert issubclass(UnsupportedFormatError, ReportError)
    assert issubclass(MissingDependencyError, ReportError)


# --------------------------------------------------------------------------- #
# Excel — depends on optional openpyxl
# --------------------------------------------------------------------------- #


@pytest.mark.skipif(_installed("openpyxl"), reason="openpyxl present — tested by the write-path case")
def test_excel_missing_dependency_raises_cleanly(tmp_path):
    with pytest.raises(MissingDependencyError):
        export_excel(SAMPLE, str(tmp_path / "r.xlsx"))
    with pytest.raises(MissingDependencyError):
        export(SAMPLE, "excel", str(tmp_path / "r2.xlsx"))


@pytest.mark.skipif(not _installed("openpyxl"), reason="openpyxl not installed")
def test_excel_writes_xlsx_when_available(tmp_path):
    path = tmp_path / "report.xlsx"
    result = export(SAMPLE, ExportFormat.EXCEL, str(path), title="Sales")
    assert result.engine == "openpyxl"
    assert result.format == "excel"
    assert result.byte_size > 0
    assert path.exists()

    # Round-trip: the header and a known numeric cell must survive.
    from openpyxl import load_workbook

    sheet = load_workbook(str(path)).active
    assert [c.value for c in sheet[1]] == ["BRANCH", "TOTAL_SALES"]
    assert sheet.cell(row=2, column=2).value == 120


# --------------------------------------------------------------------------- #
# PDF — depends on optional reportlab
# --------------------------------------------------------------------------- #


@pytest.mark.skipif(_installed("reportlab"), reason="reportlab present — tested by the write-path case")
def test_pdf_missing_dependency_raises_cleanly(tmp_path):
    with pytest.raises(MissingDependencyError):
        export_pdf(SAMPLE, str(tmp_path / "r.pdf"))
    with pytest.raises(MissingDependencyError):
        export(SAMPLE, "pdf", str(tmp_path / "r2.pdf"))


@pytest.mark.skipif(not _installed("reportlab"), reason="reportlab not installed")
def test_pdf_writes_file_when_available(tmp_path):
    path = tmp_path / "report.pdf"
    result = export(SAMPLE, ExportFormat.PDF, str(path), title="Sales")
    assert result.engine == "reportlab"
    assert result.byte_size > 0
    # A well-formed PDF begins with the %PDF- magic bytes.
    assert path.read_bytes().startswith(b"%PDF-")


def test_export_format_values():
    assert ExportFormat.CSV.value == "csv"
    assert ExportFormat.EXCEL.value == "excel"
    assert ExportFormat.PDF.value == "pdf"
